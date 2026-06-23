# Última Atualização: 26/01/2026 14:50 (Horário de Brasília)
# Criado por Reinaldo de Almeida Pinheiro em Python e A.I. Copilot

import requests
import random
import pywhatkit
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from datetime import datetime
import yfinance as yf
import os
import time
import webbrowser
import tkinter as tk
from tkinter import ttk
import pyautogui
import re

# Sequência numérica
arquivo_sequencia = "Sequencia.rpc"
ano_atual = datetime.now().year % 100
def carregar_sequencia():
    if os.path.exists(arquivo_sequencia):
        with open(arquivo_sequencia, "r", encoding="utf-8") as f:
            linha = f.read().strip()
            try:
                numero, ano = linha.split("/")
                if int(ano) == ano_atual:
                    return int(numero), ano_atual
            except:
                pass
    return 1, ano_atual
numero_boletim, ano_boletim = carregar_sequencia()
numero_formatado = f"{numero_boletim:04d}/{ano_boletim}"
titulo_boletim = f"Boletim Diário RPC - Número: {numero_formatado}"

# Frase do dia
arquivo_frases = "frases.rpc"
if not os.path.exists(arquivo_frases):
    with open(arquivo_frases, "w", encoding="utf-8") as f:
        f.write("A persistência realiza o impossível.\nA sorte favorece os audazes.\nGrandes conquistas começam com pequenos passos.")
with open(arquivo_frases, "r", encoding="utf-8") as f:
    frases = [linha.strip() for linha in f if linha.strip()]
frase_dia = random.choice(frases)

# Dia da semana
dias_semana = {
    0: "SEGUNDA-FEIRA",
    1: "TERÇA-FEIRA",
    2: "QUARTA-FEIRA",
    3: "QUINTA-FEIRA",
    4: "SEXTA-FEIRA",
    5: "SÁBADO",
    6: "DOMINGO"
}
agora = datetime.now()
dia_semana = dias_semana[agora.weekday()]

# Comemorações via arquivo
def get_comemoracoes_arquivo():
    caminho = "comemora.rpc"
    hoje = agora.strftime("%d/%m")
    comemoracoes = []
    if os.path.exists(caminho):
        with open(caminho, "r", encoding="utf-8") as f:
            for linha in f:
                if linha.strip().startswith(hoje):
                    partes = linha.strip().split(" - ", 1)
                    if len(partes) == 2:
                        comemoracoes.append(f"🎉 {partes[1]}")
    return comemoracoes or ["🎉 Nenhuma comemoração registrada."]
comemoracoes = get_comemoracoes_arquivo()

# Interface gráfica
janela = tk.Tk()
janela.title(titulo_boletim)
janela.geometry("500x250")
janela.resizable(False, False)
tk.Label(janela, text=titulo_boletim, font=("Arial", 16, "bold")).pack(pady=10)
tk.Label(janela, text="Criado em 14/10/2025 • Versão 1.1", font=("Arial", 12)).pack()
mensagem = tk.Label(janela, text="⏳ Gerando boletim...", font=("Arial", 14), fg="blue")
mensagem.pack(pady=20)
botao_fechar = ttk.Button(janela, text="Fechar", command=janela.destroy)
botao_fechar.pack(pady=10)
botao_fechar["state"] = "disabled"
janela.update()

# Saudação
hora = agora.hour
if hora < 12:
    saudacao = "Bom dia!"
elif hora < 18:
    saudacao = "Boa tarde!"
else:
    saudacao = "Boa noite!"
data_hora = agora.strftime("%d/%m/%Y")
linha = [data_hora, saudacao]

# Caminhos e API
arquivo_excel = "boletim_rpc.xlsx"
arquivo_envios = "envia.rpc"
api_key = "7be2ec2d8d42047fb862869b55d75f83"

# Planilha
if os.path.exists(arquivo_excel):
    wb = load_workbook(arquivo_excel)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Data", "Saudação", "Dólar", "Euro", "Ibovespa", "Ibovespa (%)",
        "Tempo", "Signo", "Mensagem", "Trabalho", "Cor", "Números",
        "Notícias Gerais", "Saúde", "Tecnologia", "Futebol"
    ])

# === FIM DA PARTE 1 ===
# === INÍCIO DA PARTE 2 ===

def get_cotacao_api(moeda):
    url = f"https://economia.awesomeapi.com.br/json/last/{moeda}-BRL"
    try:
        r = requests.get(url, timeout=10)
        valor = float(r.json()[f"{moeda}BRL"]["bid"])
        return f"{valor:.2f}"
    except Exception as e:
        print(f"Erro ao buscar {moeda}: {e}")
        return "Erro"
dolar = get_cotacao_api("USD")
euro = get_cotacao_api("EUR")

def get_ibovespa():
    try:
        ibov = yf.Ticker("^BVSP")
        dados = ibov.history(period="2d")
        if dados.empty or "Close" not in dados:
            return "Erro", "Erro"
        atual = round(dados["Close"].iloc[-1], 2)
        anterior = round(dados["Close"].iloc[-2], 2)
        variacao = round(((atual - anterior) / anterior) * 100, 2)
        return atual, f"{variacao:+.2f}%"
    except Exception as e:
        print("Erro Ibovespa:", e)
        return "Erro", "Erro"
ibov_valor, ibov_var = get_ibovespa()

def get_tempo(cidade="Sao Paulo", api_key=api_key):
    try:
        url = f"http://api.openweathermap.org/data/2.5/weather?q={cidade}&appid={api_key}&units=metric&lang=pt_br"
        r = requests.get(url, timeout=10)
        dados = r.json()
        temp_min = dados["main"]["temp_min"]
        temp_max = dados["main"]["temp_max"]
        chance_chuva = dados.get("clouds", {}).get("all", 0)
        return f"Mín: {temp_min:.1f}°C • Máx: {temp_max:.1f}°C • Previsão de chuva: {chance_chuva}%"
    except Exception as e:
        print("Erro tempo:", e)
        return "Erro"
tempo = get_tempo()

def get_horoscopo():
    signos = ["aries","touro","gemeos","cancer","leao","virgem","libra","escorpiao","sagitario","capricornio","aquario","peixes"]
    signo = random.choice(signos)
    url = f"https://portaldosigno.com.br/portal/horoscopo-diario/{signo}"
    try:
        response = requests.get(url, timeout=10)
        if response.status_code != 200:
            raise Exception("Site indisponível")
        soup = BeautifulSoup(response.text, 'html.parser')
        paragrafos = soup.find_all('p')
        mensagemdodia = trabalho = cor = numeros = None
        for p in paragrafos:
            texto = p.get_text().lower()
            if "mensagem do dia" in texto and not mensagemdodia:
                mensagemdodia = re.sub(r"mensagem do dia[:\-–]*", "", p.get_text(), flags=re.I).strip()
            elif "carreira" in texto and not trabalho:
                trabalho = re.sub(r"carreira[:\-–]*", "", p.get_text(), flags=re.I).strip()
            elif "cor do dia" in texto and not cor:
                cor = re.sub(r"cor do dia[:\-–]*", "", p.get_text(), flags=re.I).strip()
            elif "números do dia" in texto and not numeros:
                numeros = re.sub(r"números do dia[:\-–]*", "", p.get_text(), flags=re.I).strip()
        return {"signo": signo.capitalize(),"mensagemdodia": mensagemdodia or "Não encontrado","trabalho": trabalho or "Não encontrado","cor": cor or "Não encontrada","numeros": numeros or "Não encontrados"}
    except Exception as e:
        print("Erro horóscopo:", e)
        return {"signo": signo.capitalize(),"mensagemdodia": "Erro","trabalho": "Erro","cor": "Erro","numeros": "Erro"}
horoscopo = get_horoscopo()

def get_noticias_titulos(url, limite, icone):
    try:
        r = requests.get(url, timeout=10)
        soup = BeautifulSoup(r.text, "html.parser")
        posts = soup.find_all("a", class_="feed-post-link", limit=limite)
        return [f"{icone} {post.text.strip()}\n{post['href']}" for post in posts]
    except Exception as e:
        print("Erro ao buscar notícias:", e)
        return [f"{icone} Erro ao buscar notícias.\n#"]

noticias_gerais = get_noticias_titulos("https://g1.globo.com/", 10, "📰")
noticias_saude = get_noticias_titulos("https://g1.globo.com/saude/", 5, "⚕️")
noticias_tech = get_noticias_titulos("https://g1.globo.com/economia/tecnologia/", 5, "💻")
noticias_futebol = get_noticias_titulos("https://ge.globo.com/futebol/", 5, "⚽")

linha.extend([f"R$ {dolar}", f"R$ {euro}", ibov_valor, ibov_var, tempo])
linha.extend([
    horoscopo["signo"], horoscopo["mensagemdodia"], horoscopo["trabalho"],
    horoscopo["cor"], horoscopo["numeros"]
])
linha.append("\n".join(noticias_gerais))
linha.append("\n".join(noticias_saude))
linha.append("\n".join(noticias_tech))
linha.append("\n".join(noticias_futebol))

ws.append(linha)
wb.save(arquivo_excel)

# === FIM DA PARTE 2 ===
# === INÍCIO DA PARTE 3 ===

# Hora atual
hora_atual = agora.strftime("%H:%M")
data_hora_completa = f"{linha[0]} {hora_atual}"
titulo_boletim_whatsapp = f"📊 {titulo_boletim} — {data_hora_completa}"

# Boletim para WhatsApp
boletim = f"""
{titulo_boletim_whatsapp}
📝 Frase do dia: {frase_dia}
**{linha[1].upper()}**
{linha[0]}
**{dia_semana}**

{"\n".join(comemoracoes)}

💰 Economia
• Dólar: {linha[2]}
• Euro: {linha[3]}
• Ibovespa: {linha[4]} pts ({linha[5]})

🌦️ Tempo em São Paulo
• {linha[6]}

🔮 Horóscopo — {horoscopo['signo']}
❤️ {horoscopo['mensagemdodia']}
💼 {horoscopo['trabalho']}
🎨 {horoscopo['cor']}
🎲 {horoscopo['numeros']}

📰 Notícias do dia:
(Para ver os links da notícia veja no final o site!)
"""

for lista in [noticias_gerais, noticias_saude, noticias_tech, noticias_futebol]:
    for noticia in lista:
        partes = noticia.split("\n")
        titulo = partes[0]
        boletim += f"{titulo}\n"

boletim += """
Para ver esse boletim na Web acesse: https://reinaldopinheiro.com.br/boletim.html
Gostou! Ajude a manter esse projeto!
Doação chave pix: doe@reinaldopinheiro.com.br

📚 Leia o livro: Conhecendo a área de T.I. de Reinaldo de Almeida Pinheiro
http://reinaldopinheiro.com.br/livro.html
"""

# Envio via WhatsApp
try:
    with open(arquivo_envios, "r", encoding="utf-8") as f:
        linhas_envio = [linha.strip() for linha in f if linha.strip()]
    tempo_espera = int(linhas_envio[0])
    numeros = linhas_envio[1:]

    if tempo_espera == 999:
        print("⏸️ Envio via WhatsApp desativado (tempo = 999)")
    else:
        if tempo_espera > 0:
            print(f"⏳ Aguardando {tempo_espera} minuto(s)...")
            time.sleep(tempo_espera * 60)
        else:
            print("🚀 Envio imediato iniciado...")

        for numero in numeros:
            pywhatkit.sendwhatmsg_instantly(numero, boletim, wait_time=20, tab_close=True)
            print(f"✅ Boletim enviado para {numero}")
except Exception as e:
    print("❌ Erro ao enviar boletins:", e)

# Geração do HTML
html_boletim = f"""<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <title>{titulo_boletim} — {data_hora_completa}</title>
    <style>
        body {{ font-family: Arial, sans-serif; padding: 20px; background-color: #f9f9f9; }}
        h1 {{ font-size: 28px; color: #2c3e50; }}
        .data {{ font-size: 22px; }}
        .saudacao {{ font-size: 18px; font-weight: bold; text-transform: uppercase; }}
        ul {{ padding-left: 20px; }}
        li {{ margin-bottom: 10px; }}
    </style>
</head>
<body>
    <h1>{titulo_boletim} — {data_hora_completa}</h1>
    <p><strong>Frase do dia:</strong> {frase_dia}</p>
    <div class="saudacao">{linha[1].upper()}</div>
    <div class="data">{linha[0]}</div>
    <div class="saudacao">{dia_semana}</div>
    <ul>
"""

for item in comemoracoes:
    html_boletim += f"<li>{item}</li>\n"

html_boletim += f"""
    </ul>
    <h2>💰 Economia</h2>
    <ul>
        <li>Dólar: {linha[2]}</li>
        <li>Euro: {linha[3]}</li>
        <li>Ibovespa: {linha[4]} pts ({linha[5]})</li>
    </ul>
    <h2>🌦️ Tempo em São Paulo</h2>
    <ul>
        <li>{linha[6]}</li>
    </ul>
    <h2>🔮 Horóscopo — {horoscopo['signo']}</h2>
    <ul>
        <li>❤️ {horoscopo['mensagemdodia']}</li>
        <li>💼 {horoscopo['trabalho']}</li>
        <li>🎨 {horoscopo['cor']}</li>
        <li>🎲 {horoscopo['numeros']}</li>
    </ul>
    <h2>📰 Notícias do dia (Fonte: <a href="https://g1.globo.com" target="_blank">g1.globo.com</a>)</h2>
    <ul>
"""

def add_noticias_html(lista):
    html = ""
    for noticia in lista:
        partes = noticia.split("\n")
        if len(partes) == 2:
            titulo, link = partes
            html += f'<li>{titulo} — <a href="{link}" target="_blank">Link</a></li>\n'
        else:
            html += f"<li>{noticia}</li>\n"
    return html

html_boletim += add_noticias_html(noticias_gerais)
html_boletim += add_noticias_html(noticias_saude)
html_boletim += add_noticias_html(noticias_tech)
html_boletim += add_noticias_html(noticias_futebol)

html_boletim += """
    </ul>
    <p><strong>Ajude a manter esse projeto.</strong></p>
    <footer>
        <p><strong>Doação chave pix:</strong> <em>doe@reinaldopinheiro.com.br</em></p>
        <p>📚 Leia o livro: <em>Conhecendo a área de T.I.</em> de Reinaldo de Almeida Pinheiro —
        <a href="http://reinaldopinheiro.com.br/livro.html" target="_blank">Clique aqui</a></p>
    </footer>
</body>
</html>
"""

with open("boletim.html", "w", encoding="utf-8") as f:
    f.write(html_boletim)

# Cópia com nome AAMMDD_Boletim.html
nome_copia = agora.strftime("%y%m%d") + "_Boletim.html"
with open(nome_copia, "w", encoding="utf-8") as f:
    f.write(html_boletim)

# Atualiza sequência
proximo_numero = numero_boletim + 1
nova_linha = f"{proximo_numero:04d}/{ano_boletim}"
with open(arquivo_sequencia, "w", encoding="utf-8") as f:
    f.write(nova_linha)

# Temporizador de 1 minuto
def fechar_automaticamente(segundos=60):
    def atualizar():
        nonlocal segundos
        if segundos > 0:
            botao_fechar.config(text=f"Fechar ({segundos}s)")
            segundos -= 1
            janela.after(1000, atualizar)
        else:
            janela.destroy()
    atualizar()

mensagem.config(text="✅ Terminado com sucesso!", fg="green")
botao_fechar["state"] = "normal"
fechar_automaticamente()
janela.mainloop()

print("*** fim do programa ***")

# === FIM DA PARTE 3 ===
